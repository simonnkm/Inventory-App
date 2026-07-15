from fastapi import FastAPI, Depends, HTTPException, Query, Response, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from database import Base, engine, SessionLocal
from datetime import datetime, timedelta, timezone, date
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import jwt
from io import BytesIO
from fastapi.responses import StreamingResponse, FileResponse
from openpyxl import Workbook, load_workbook
import os
import tempfile
import models
import schemas
import crud
import auth
import re
from pathlib import Path
from uuid import uuid4

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://10.3.20.141:3000",
        "https://inv-app-omega.vercel.app",
    ],
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ORDER_COLUMNS = [
    ("Date", "order_date"),
    ("Order Placed by", "order_placed_by"),
    ("PO Number", "po_number"),
    ("Vendor", "vendor"),
    ("Category", "category"),
    ("Catalog No.", "catalog_no"),
    ("Item Name", "item_name"),
    ("# of Units", "units_ordered"),
    ("Price / Unit", "price_per_unit"),
    ("Total Price", "total_price"),
    ("Final Price (with tax & freight)", "final_price"),
    ("Availability", "availability"),
    ("Expected Delivery Date", "expected_delivery_date"),
    ("Order #", "order_number"),
    ("Delivery Date", "delivery_date"),
    ("Status", "status"),
    ("Received by", "received_by"),
    ("Date paid", "date_paid"),
    ("Amount Paid", "amount_paid"),
    ("CC/Invoice?", "cc_invoice"),
]


HEADER_ALIASES = {
    "date": "order_date",
    "order placed by": "order_placed_by",
    "po number": "po_number",
    "vendor": "vendor",
    "category": "category",
    "catalog no.": "catalog_no",
    "catalog no": "catalog_no",
    "catalog number": "catalog_no",
    "item name": "item_name",
    "# of units": "units_ordered",
    "units ordered": "units_ordered",
    "price / unit": "price_per_unit",
    "price per unit": "price_per_unit",
    "total price": "total_price",
    "final price (with tax & freight)": "final_price",
    "final price": "final_price",
    "availability": "availability",
    "expected delivery date": "expected_delivery_date",
    "order #": "order_number",
    "order number": "order_number",
    "delivery date": "delivery_date",
    "status": "status",
    "received by": "received_by",
    "dated paid": "date_paid",
    "date paid": "date_paid",
    "amount paid": "amount_paid",
    "cc/invoice?": "cc_invoice",
    "cc invoice": "cc_invoice",
}

if os.getenv("VERCEL"):
    UPLOAD_ROOT = Path(tempfile.gettempdir()) / "order_documents"
else:
    UPLOAD_ROOT = Path("uploads/order_documents")

UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

def normalize_header(value):
    return str(value or "").strip().lower()

def excel_rows_for_ai(sheet):
    rows = []

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = []

        for col_index, value in enumerate(row, start=1):
            if value not in (None, ""):
                values.append({
                    "column": col_index,
                    "value": str(value),
                })

        if values:
            rows.append({
                "row_number": row_number,
                "values": values,
            })

    return rows

def parse_date(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    # Excel serial date support, e.g. 46210
    if isinstance(value, (int, float)):
        if 20000 <= value <= 60000:
            return date(1899, 12, 30) + timedelta(days=int(value))
        return None

    text = str(value).strip()

    for fmt in ("%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return None

def strip_list_prefix(text: str, index: int | None = None):
    text = str(text or "").strip()

    if not text:
        return ""

    # Handles: 1. ABC, 1) ABC, 1 ABC
    text = re.sub(r"^\s*\d+\s*[\.\)]\s*", "", text)
    text = re.sub(r"^\s*\d+\s+", "", text)

    # Handles glued numbering like 1FEREP0402, 2FERR0192
    # Avoids touching normal catalog numbers unless the row clearly looks listed.
    if index is not None:
        expected = str(index + 1)

        if text.startswith(expected) and len(text) > len(expected):
            rest = text[len(expected):]

            if rest and rest[0].isalpha():
                text = rest.strip()

    return text.strip(" -;\t")


def split_multivalue_cell(value):
    if value in (None, ""):
        return []

    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()

    if not text:
        return []

    parts = [part.strip() for part in text.split("\n") if part and str(part).strip()]

    cleaned = []

    for index, part in enumerate(parts):
        item = strip_list_prefix(part, index)

        if item:
            cleaned.append(item)

    return cleaned


def split_units(value, count: int):
    if count <= 1:
        parsed = parse_int(value)
        return [parsed]

    if value in (None, ""):
        return [None] * count

    text = str(value).strip()

    # Handles "1\n2\n1"
    if "\n" in text:
        parts = split_multivalue_cell(text)
        nums = [parse_int(part) for part in parts]

        if len(nums) == count:
            return nums

    # Handles exported collapsed values like "121" for 3 rows -> [1, 2, 1]
    if text.isdigit() and len(text) == count and count > 1:
        return [int(char) for char in text]

    # Avoid duplicating aggregate unit counts across every split row.
    return [None] * count


def expand_order_payload(raw_payload: dict, raw_units_value):
    catalog_parts = split_multivalue_cell(raw_payload.get("catalog_no"))
    item_parts = split_multivalue_cell(raw_payload.get("item_name"))

    # For this app, rows without catalog numbers are not useful for inventory.
    if not catalog_parts:
        return []

    count = len(catalog_parts)
    unit_parts = split_units(raw_units_value, count)

    expanded = []

    for index, catalog_no in enumerate(catalog_parts):
        item_name = raw_payload.get("item_name")

        if len(item_parts) == count:
            item_name = item_parts[index]

        item_name = clean_string(item_name)

        if not item_name:
            continue

        child = dict(raw_payload)
        child["catalog_no"] = catalog_no
        child["item_name"] = item_name
        child["units_ordered"] = unit_parts[index] if index < len(unit_parts) else None

        # If a row has multiple products but only one aggregate price,
        # keep price on first child only so exports do not multiply totals.
        if count > 1 and index > 0:
            child["price_per_unit"] = None
            child["total_price"] = None
            child["final_price"] = None
            child["amount_paid"] = None

        expanded.append(child)

    return expanded

def parse_float(value):
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().lower()

    empty_values = {
        "na",
        "n/a",
        "none",
        "null",
        "varies",
        "variable",
        "tbd",
        "unknown",
        "pending",
        "-",
        "--",
    }

    if text in empty_values:
        return None

    text = (
        text.replace("$", "")
        .replace(",", "")
        .replace("usd", "")
        .strip()
    )

    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value):
    number = parse_float(value)

    if number is None:
        return None

    return int(number)

def is_service_order(payload: dict):
    combined = " ".join(
        str(value or "").lower()
        for value in [
            payload.get("vendor"),
            payload.get("category"),
            payload.get("catalog_no"),
            payload.get("item_name"),
            payload.get("po_number"),
            payload.get("order_number"),
        ]
    )

    service_keywords = [
        "service",
        "services",
        "software",
        "subscription",
        "license",
        "licence",
        "monthly",
        "annual",
        "microsoft",
        "cloud",
        "hosting",
        "internet",
        "it support",
    ]

    return any(keyword in combined for keyword in service_keywords)

def clean_string(value):
    if value in (None, ""):
        return None

    return str(value).strip()


def order_to_excel_row(order):
    return [
        order.order_date,
        order.order_placed_by,
        order.po_number,
        order.vendor,
        order.category,
        order.catalog_no,
        order.item_name,
        order.units_ordered,
        order.price_per_unit,
        order.total_price,
        order.final_price,
        order.availability,
        order.expected_delivery_date,
        order.order_number,
        order.delivery_date,
        order.status,
        order.received_by,
        order.date_paid,
        order.amount_paid,
        order.cc_invoice,
    ]

models.Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.post("/items/", response_model=schemas.ItemResponse)
def create_item(
    item: schemas.ItemCreate,
    db: Session = Depends(get_db),
    current_user = Depends(auth.require_role("admin")),
):
    created_item = crud.create_item(db, item)

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="CREATE_ITEM",
        item_id=created_item.id,
        details=f"Created item {created_item.item_name}",
    )

    return created_item

@app.post("/items/{item_id}/use", response_model=schemas.ItemResponse)
def use_item(
    item_id: str,
    amount: int = Query(..., gt = 0),
    db: Session = Depends(get_db),
    current_user = Depends(auth.require_role("admin", "user")),
):
    item = crud.use_item(db, item_id, amount)

    if item is None:
        raise HTTPException(status_code=404, detail="Item not found")

    if item == "invalid_amount":
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")

    if item == "not_enough_quantity":
        raise HTTPException(status_code=400, detail="Not enough quantity")

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="USE_ITEM",
        item_id=item.id,
        details=f"Used {amount} of {item.item_name}",
    )

    return item

@app.post("/items/{item_id}/restock", response_model=schemas.ItemResponse)
def restock_item(
    item_id: str,
    amount: int = Query(..., gt = 0),
    db: Session = Depends(get_db),
    current_user = Depends(auth.require_role("admin")),
):
    item = crud.restock_item(db, item_id, amount)

    if item is None:
        raise HTTPException(status_code=404, detail="Item not found")

    if item == "invalid_amount":
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="RESTOCK_ITEM",
        item_id=item.id,
        details=f"Restocked {amount} of {item.item_name}",
    )

    return item

@app.put("/items/{item_id}", response_model = schemas.ItemResponse)
def update_item(
    item_id: str,
    item_data: schemas.ItemUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(auth.require_role("admin")),
):
    item = crud.update_item(db, item_id, item_data)
    if item is None:
        raise HTTPException(status_code=404, detail = "Item not found")
    
    crud.create_audit_log(
        db,
        username = current_user.username,
        action = "UPDATE_ITEM",
        item_id = item.id,
        details = f"Updated item {item.item_name}",
    )

    return item

@app.get("/items/", response_model=list[schemas.ItemResponse])
def get_items(
    name: str | None = None,
    storage_id: str | None = None,
    expiring_before: date | None = None,
    status: str | None = None,
    category: str | None = None,
    shelf_num: str | None = None,
    db: Session = Depends(get_db),
    current_user = Depends(auth.get_current_user),
):
    return crud.get_items_filtered(
        db,
        name=name,
        storage_id=storage_id,
        expiring_before=expiring_before,
        status = status
    )

@app.get("/items/{item_id}", response_model=schemas.ItemResponse)
def get_item(item_id: str, db: Session = Depends(get_db), current_user = Depends(auth.get_current_user)):
    item = crud.get_item(db, item_id)

    if item is None:
        raise HTTPException(status_code = 404, detail = "Item not found")
    
    return item

@app.delete("/items/{item_id}", response_model=schemas.ItemResponse)
def delete_item(item_id: str, db: Session = Depends(get_db), current_user = Depends(auth.require_role("admin")),):
    item = crud.delete_item(db, item_id)

    if item is None:
        raise HTTPException(status_code = 404, detail = "Item not found")
    
    crud.create_audit_log(
        db,
        username = current_user.username,
        action = "DELETE_ITEM",
        item_id = item.id,
        details = f"Deleted item {item.item_name}",
    )

    return item

@app.get("/item-types/", response_model=list[schemas.ItemTypeResponse])
def get_item_types(
    db: Session = Depends(get_db),
    current_user=Depends(auth.get_current_user),
):
    return crud.get_item_types(db)


@app.post("/item-types/", response_model=schemas.ItemTypeResponse)
def create_item_type(
    payload: schemas.ItemTypeCreate,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    item_type = crud.create_item_type(db, payload)

    return {
        "id": item_type.id,
        "name": item_type.name,
        "category": item_type.category,
        "brand": item_type.brand,
        "reorder_threshold": item_type.reorder_threshold,
        "critical_threshold": item_type.critical_threshold,
        "notes": item_type.notes,
        "total_quantity": 0,
    }

@app.post("/register", response_model=schemas.UserResponse)
def register(user: schemas.UserCreate, db: Session = Depends(get_db)):
    user_count = db.query(models.User).count()

    if user_count > 0:
        raise HTTPException(
            status_code=403,
            detail="Registration disabled. Ask an admin to create your account.",
        )

    user.role = "admin"

    return crud.create_user(db, user)

@app.post("/users/", response_model=schemas.UserResponse)
def create_user_by_admin(
    user: schemas.UserCreate,
    db: Session = Depends(get_db),
    current_user = Depends(auth.require_role("admin")),
):
    existing = crud.get_user(db, user.username)

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Username already exists",
        )

    return crud.create_user(db, user)

@app.get("/users/", response_model=list[schemas.UserResponse])
def list_users(
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    return crud.get_users(db)

@app.delete("/users/{user_id}", status_code=204)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    if current_user.id == user_id:
        raise HTTPException(
            status_code=400,
            detail="You cannot delete your own account while logged in",
        )

    deleted_user = crud.delete_user(db, user_id)

    if deleted_user is None:
        raise HTTPException(
            status_code=404,
            detail="User not found",
        )

    return Response(status_code=204)

@app.post("/login")
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):

    auth_user = crud.authenticate_user(db, form_data.username, form_data.password)

    if not auth_user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    token = auth.create_access_token(
        data = {"sub": auth_user.username}
    )

    return {"access_token": token, "token_type": "bearer",}

@app.get("/audit-logs/", response_model= list[schemas.AuditLogResponse])
def get_audit_logs(
    db: Session = Depends(get_db),
    current_user = Depends(auth.require_role("admin")),
):
    return crud.get_audit_logs(db)

@app.get("/items/low-stock/", response_model=list[schemas.ItemResponse])
def get_low_stock_items(
    db: Session = Depends(get_db),
    current_user = Depends(auth.get_current_user),
):
    return crud.get_low_stock_items(db)

@app.get("/me", response_model=schemas.UserResponse)
def read_me(current_user = Depends(auth.get_current_user)):
    return current_user

@app.get("/dashboard", response_model=schemas.DashBoardStats)
def dashboard(
    db: Session = Depends(get_db),
    current_user = Depends(auth.get_current_user),
):
    return crud.get_dashboard_stats(db)

@app.post("/items/{item_id}/transaction", response_model=schemas.ItemResponse)
def create_transaction(
    item_id: str,
    change_amount: int = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin", "user")),
):
    if change_amount == 0:
        raise HTTPException(
            status_code=400,
            detail="Transaction amount cannot be zero",
        )

    if change_amount > 0 and current_user.role != "admin":
        raise HTTPException(
            status_code=403,
            detail="Only admins can increase quantity",
        )

    result = crud.create_transaction(db, item_id, change_amount)

    if result is None:
        raise HTTPException(status_code=404, detail="Item not found")

    if result == "not_enough_quantity":
        raise HTTPException(
            status_code=400,
            detail="Not enough quantity available",
        )

    item, old_quantity, new_quantity = result

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="TRANSACTION",
        item_id=item.id,
        details=f"Changed {item.item_name} by {change_amount}",
        old_quantity=old_quantity,
        change_amount=change_amount,
        new_quantity=new_quantity,
    )

    return item

@app.get("/orders/", response_model=list[schemas.OrderResponse])
def list_orders(
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    return crud.get_orders(db)

@app.post("/orders/", response_model=schemas.OrderResponse)
def create_order(
    order: schemas.OrderCreate,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    created_order = crud.create_order(db, order)

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="CREATE_ORDER",
        item_id=None,
        details=(
            f"Created order #{created_order.id}: "
            f"{created_order.item_name} "
            f"({created_order.catalog_no}) from {created_order.vendor or 'unknown vendor'}"
        ),
    )

    return created_order


@app.post("/orders/import", response_model=list[schemas.OrderResponse])
async def import_orders(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    filename = file.filename or ""

    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Please upload an .xlsx file",
        )

    content = await file.read()
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = rows[0]
    field_indexes = {}

    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        field_name = HEADER_ALIASES.get(normalized)

        if field_name:
            field_indexes[field_name] = index

    if "item_name" not in field_indexes:
        raise HTTPException(
            status_code=400,
            detail="Excel file must include an Item Name column",
        )

    created_orders = []

    for row in rows[1:]:
        if all(value in (None, "") for value in row):
            continue

        def value_for(field_name):
            index = field_indexes.get(field_name)

            if index is None or index >= len(row):
                return None

            return row[index]
        item_name = clean_string(value_for("item_name"))

        if not item_name:
            continue
        
        order_placed_by = clean_string(value_for("order_placed_by"))
        vendor = clean_string(value_for("vendor"))
        category = clean_string(value_for("category"))
        catalog_no = clean_string(value_for("catalog_no"))

        has_real_order_identifier = any(
            [
                order_placed_by,
                vendor,
                category,
                catalog_no,
            ]
        )

        if not has_real_order_identifier:
            continue

        raw_payload = {
            "order_date": parse_date(value_for("order_date")),
            "order_placed_by": order_placed_by,
            "po_number": clean_string(value_for("po_number")),
            "vendor": vendor,
            "category": category,
            "catalog_no": catalog_no,
            "item_name": item_name,
            "units_ordered": parse_int(value_for("units_ordered")),
            "price_per_unit": parse_float(value_for("price_per_unit")),
            "total_price": parse_float(value_for("total_price")),
            "final_price": parse_float(value_for("final_price")),
            "availability": clean_string(value_for("availability")),
            "expected_delivery_date": parse_date(value_for("expected_delivery_date")),
            "order_number": clean_string(value_for("order_number")),
            "delivery_date": parse_date(value_for("delivery_date")),
            "status": clean_string(value_for("status")) or "Ordered",
            "received_by": clean_string(value_for("received_by")),
            "date_paid": parse_date(value_for("date_paid")),
            "amount_paid": parse_float(value_for("amount_paid")),
            "cc_invoice": clean_string(value_for("cc_invoice")),
        }

        if is_service_order(raw_payload):
            continue

        expanded_payloads = expand_order_payload(
            raw_payload,
            value_for("units_ordered"),
        )

        for expanded_payload in expanded_payloads:
            if is_service_order(expanded_payload):
                continue

            payload = schemas.OrderCreate(**expanded_payload)

            created_orders.append(crud.create_order(db, payload))

    if created_orders:
        crud.create_audit_log(
            db,
            username=current_user.username,
            action="IMPORT_ORDERS",
            item_id=None,
            details=f"Imported {len(created_orders)} orders from {filename}",
        )
    
    return created_orders


@app.get("/orders/export")
def export_orders(
    ids: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    if ids:
        try:
            order_ids = [int(value) for value in ids.split(",") if value.strip()]
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="ids must be comma-separated integers",
            )

        orders = crud.get_orders_by_ids(db, order_ids)
    else:
        orders = crud.get_orders(db)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"

    sheet.append([label for label, _field in ORDER_COLUMNS])

    for order in orders:
        sheet.append(order_to_excel_row(order))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="orders.xlsx"'
        },
    )

@app.get(
    "/orders/{order_id}/documents",
    response_model=list[schemas.OrderDocumentResponse],
)
def list_order_documents(
    order_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    order = crud.get_order_by_id(db, order_id)

    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")

    return crud.get_order_documents(db, order_id)

@app.delete("/order-documents/{document_id}")
def delete_order_document(
    document_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    document = crud.get_order_document(db, document_id)

    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")

    file_path = Path(document.file_path) if document.file_path else None
    original_filename = document.original_filename
    order_id = document.order_id

    crud.delete_order_document(db, document_id)

    if file_path and file_path.exists():
        file_path.unlink()

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="DELETE_ORDER_DOCUMENT",
        item_id=None,
        details=f"Deleted document {original_filename or document_id} from order #{order_id}",
    )

    return {"message": "Document deleted"}

@app.post(
    "/orders/{order_id}/documents",
    response_model=schemas.OrderDocumentResponse,
)
async def upload_order_document(
    order_id: int,
    document_type: str = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    order = crud.get_order_by_id(db, order_id)

    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")

    allowed_types = {
        "confirmation",
        "invoice",
        "delivery",
        "packing_slip",
        "shipping",
        "other",
    }

    if document_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail=f"document_type must be one of: {', '.join(sorted(allowed_types))}",
        )

    original_filename = file.filename or "upload"
    suffix = Path(original_filename).suffix
    stored_filename = f"{uuid4().hex}{suffix}"
    file_path = UPLOAD_ROOT / stored_filename

    content = await file.read()
    file_path.write_bytes(content)

    document = crud.create_order_document(
        db,
        order_id=order_id,
        document_type=document_type,
        source="manual",
        original_filename=original_filename,
        stored_filename=stored_filename,
        file_path=str(file_path),
        content_type=file.content_type,
        reviewed=True,
    )

    crud.create_order_event(
        db,
        order_id=order_id,
        event_type="DOCUMENT_UPLOADED",
        notes=f"Uploaded {document_type}: {original_filename}",
        created_by=current_user.username,
    )

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="UPLOAD_ORDER_DOCUMENT",
        item_id=None,
        details=(
            f"Uploaded {document_type} document for order #{order_id}: "
            f"{original_filename}"
        ),
    )

    if document_type == "confirmation":
        crud.mark_order_confirmed(db, order_id, current_user.username)

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="ORDER_CONFIRMED",
        item_id=None,
        details=f"Order #{order_id} marked confirmed from uploaded confirmation",
    )

    if document_type == "invoice":
        crud.mark_order_invoice_received(db, order_id, current_user.username)

    if document_type in {"delivery", "packing_slip"}:
        crud.mark_order_delivered(
            db,
            order_id=order_id,
            delivery_date_value=date.today(),
            received_by=current_user.username,
            username=current_user.username,
            notes=f"{document_type} uploaded",
        )

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="ORDER_DELIVERED",
        item_id=None,
        details=f"Order #{order_id} marked delivered from uploaded {document_type}",
    )
    return document


@app.get(
    "/orders/{order_id}/events",
    response_model=list[schemas.OrderEventResponse],
)
def list_order_events(
    order_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    order = crud.get_order_by_id(db, order_id)

    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")

    return crud.get_order_events(db, order_id)


@app.post("/orders/{order_id}/mark-delivered", response_model=schemas.OrderResponse)
def mark_order_delivered(
    order_id: int,
    payload: schemas.MarkDeliveredRequest,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    order = crud.mark_order_delivered(
        db,
        order_id=order_id,
        delivery_date_value=payload.delivery_date,
        received_by=payload.received_by,
        username=current_user.username,
        notes=payload.notes,
    )

    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")
    
    crud.create_audit_log(
        db,
        username=current_user.username,
        action="ORDER_DELIVERED",
        item_id=None,
        details=(
            f"Marked order #{order_id} delivered. "
            f"Inventory updated for {order.item_name} ({order.catalog_no})"
        ),
    )
    return order


@app.post("/orders/{order_id}/mark-paid", response_model=schemas.OrderResponse)
def mark_order_paid(
    order_id: int,
    payload: schemas.MarkPaidRequest,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    order = crud.mark_order_paid(
        db,
        order_id=order_id,
        date_paid_value=payload.date_paid,
        amount_paid=payload.amount_paid,
        cc_invoice=payload.cc_invoice,
        username=current_user.username,
        notes=payload.notes,
    )

    if order is None:
        raise HTTPException(status_code=404, detail="Order not found")

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="ORDER_PAID",
        item_id=None,
        details=(
            f"Marked order #{order_id} paid. "
            f"Amount paid: {order.amount_paid if order.amount_paid is not None else 'unknown'}"
        ),
    )

    return order

@app.get(
    "/orders/{order_id}/documents",
    response_model=list[schemas.OrderDocumentResponse],
)
def get_order_documents(
    order_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    return crud.get_order_documents(db, order_id)

@app.get("/order-documents/{document_id}/download")
def download_order_document(
    document_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    document = crud.get_order_document(db, document_id)

    if document is None:
        raise HTTPException(status_code=404, detail="Document not found")

    if not document.file_path:
        raise HTTPException(status_code=404, detail="Document path missing")

    file_path = Path(document.file_path)

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File missing from server")

    return FileResponse(
        path=file_path,
        media_type=document.content_type or "application/octet-stream",
        filename=document.original_filename or document.stored_filename,
    )

@app.get(
    "/items/{item_id}/comments",
    response_model=list[schemas.ItemCommentResponse],
)
def get_item_comments(
    item_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(auth.get_current_user),
):
    return crud.get_item_comments(db, item_id)


@app.post(
    "/items/{item_id}/comments",
    response_model=schemas.ItemCommentResponse,
)
def create_item_comment(
    item_id: str,
    payload: schemas.ItemCommentCreate,
    db: Session = Depends(get_db),
    current_user=Depends(auth.get_current_user),
):
    if not payload.comment.strip():
        raise HTTPException(status_code=400, detail="Comment cannot be empty")

    comment = crud.create_item_comment(
        db,
        item_id=item_id,
        username=current_user.username,
        comment=payload.comment,
    )

    if comment is None:
        raise HTTPException(status_code=404, detail="Item not found")

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="ITEM_COMMENT",
        item_id=None,
        details=f"Commented on item {item_id}: {payload.comment[:100]}",
    )

    return comment


@app.delete("/item-comments/{comment_id}")
def delete_item_comment(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(auth.require_role("admin")),
):
    comment = crud.delete_item_comment(db, comment_id)

    if comment is None:
        raise HTTPException(status_code=404, detail="Comment not found")

    crud.create_audit_log(
        db,
        username=current_user.username,
        action="DELETE_ITEM_COMMENT",
        item_id=None,
        details=f"Deleted comment #{comment_id} from item {comment.item_id}",
    )

    return {"message": "Comment deleted"}